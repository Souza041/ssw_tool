import csv
import re
import unicodedata

from pathlib import Path

from openpyxl import load_workbook

CLIENTE_COLUNA = "Cliente Pagador"
CIDADE_COLUNA = "Cidade do Destinatario"
UNIDADE_COLUNA = "Unidade Emissora"
CTRC_COLUNA = "Serie/Numero CTRC"


def normalizar_texto(valor: object) -> str:
    texto = str(valor or "")

    texto = texto.replace("\xa0", " ")
    texto = " ".join(texto.split())
    texto = texto.strip().upper()

    return texto


def normalizar_sem_acento(valor: object) -> str:
    texto = normalizar_texto(valor)

    return "".join(
        caractere
        for caractere in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caractere) != "Mn"
    )


def detectar_encoding(arquivo: Path) -> str:
    conteudo = arquivo.read_bytes()

    # UTF-8 com BOM
    if conteudo.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"

    # UTF-16 LE com BOM
    if conteudo.startswith(b"\xff\xfe"):
        return "utf-16-le"

    # UTF-16 BE com BOM
    if conteudo.startswith(b"\xfe\xff"):
        return "utf-16-be"

    # Sem BOM, não tentamos UTF-16 automaticamente.
    for encoding in (
        "utf-8",
        "cp1252",
        "latin1",
    ):
        try:
            conteudo.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue

    return "latin1"


def localizar_cabecalho(
    linhas: list[list[str]],
) -> int:
    obrigatorias = {
        CLIENTE_COLUNA,
        CIDADE_COLUNA,
        UNIDADE_COLUNA,
        CTRC_COLUNA,
    }

    obrigatorias_normalizadas = {
        normalizar_sem_acento(nome)
        for nome in obrigatorias
    }

    for indice, linha in enumerate(
        linhas[:50]
    ):
        nomes = {
            normalizar_sem_acento(coluna)
            for coluna in linha
            if normalizar_texto(coluna)
        }

        if obrigatorias_normalizadas.issubset(
            nomes
        ):
            return indice

    primeiras_linhas = []

    for indice, linha in enumerate(
        linhas[:10]
    ):
        valores = [
            normalizar_texto(valor)
            for valor in linha
            if normalizar_texto(valor)
        ]

        primeiras_linhas.append(
            f"Linha {indice + 1}: {valores[:10]}"
        )

    raise ValueError(
        "Não foi possível localizar o cabeçalho "
        "do relatório OP455.\n"
        + "\n".join(primeiras_linhas)
    )

def carregar_linhas_xlsx(
    arquivo: Path,
) -> list[list[str]]:
    workbook = load_workbook(
        filename=arquivo,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active

        linhas = []

        for row in worksheet.iter_rows(
            values_only=True,
        ):
            linha = [
                "" if valor is None else str(valor)
                for valor in row
            ]

            linhas.append(linha)

        return linhas
    finally:
        workbook.close()

def carregar_relatorio(
    arquivo: Path,
) -> list[dict]:
    arquivo = Path(arquivo)

    if not arquivo.exists():
        raise FileNotFoundError(
            f"Arquivo não encontrado: {arquivo}"
        )

    extensao = arquivo.suffix.lower()

    if extensao in {".xlsx", ".xlsm"}:
        linhas = carregar_linhas_xlsx(arquivo)

    elif extensao in {
        ".sswweb",
        ".csv",
        ".txt",
    }:
        encoding = detectar_encoding(arquivo)

        try:
            texto = arquivo.read_text(
                encoding=encoding,
            )
        except UnicodeError:
            texto = arquivo.read_text(
                encoding="latin1",
                errors="replace",
            )

        amostra = texto[:5000]

        try:
            dialect = csv.Sniffer().sniff(
                amostra,
                delimiters=";,\t",
            )
            delimitador = dialect.delimiter
        except csv.Error:
            delimitador = ";"

        linhas = list(
            csv.reader(
                texto.splitlines(),
                delimiter=delimitador,
            )
        )

    else:
        raise ValueError(
            "Formato de arquivo não suportado: "
            f"{extensao or 'sem extensão'}"
        )

    indice_cabecalho = localizar_cabecalho(
        linhas
    )

    cabecalho = [
        normalizar_texto(coluna)
        for coluna in linhas[indice_cabecalho]
    ]

    registros = []

    for linha in linhas[indice_cabecalho + 1:]:
        if not any(
            normalizar_texto(valor)
            for valor in linha
        ):
            continue

        if len(linha) < len(cabecalho):
            linha += [""] * (
                len(cabecalho) - len(linha)
            )

        registro = {
            cabecalho[indice]: linha[indice]
            for indice in range(len(cabecalho))
        }

        registros.append(registro)

    return registros

def encontrar_coluna(
    registro: dict,
    nome_esperado: str,
) -> str:
    esperado = normalizar_sem_acento(nome_esperado)

    for coluna in registro:
        if normalizar_sem_acento(coluna) == esperado:
            return coluna

    raise KeyError(
        f"Coluna não encontrada no relatório: {nome_esperado}"
    )


def filtrar_registros(
    registros: list[dict],
    clientes_permitidos: set[str],
    cidades_permitidas: set[str],
    unidade_emissora: str,
) -> list[dict]:
    if not registros:
        return []

    exemplo = registros[0]

    coluna_cliente = encontrar_coluna(
        exemplo,
        CLIENTE_COLUNA,
    )
    coluna_cidade = encontrar_coluna(
        exemplo,
        CIDADE_COLUNA,
    )
    coluna_unidade = encontrar_coluna(
        exemplo,
        UNIDADE_COLUNA,
    )
    coluna_ctrc = encontrar_coluna(
        exemplo,
        CTRC_COLUNA,
    )

    clientes_normalizados = {
        normalizar_sem_acento(cliente)
        for cliente in clientes_permitidos
    }

    cidades_normalizadas = {
        normalizar_sem_acento(cidade)
        for cidade in cidades_permitidas
    }

    unidade_normalizada = normalizar_sem_acento(
        unidade_emissora
    )

    filtrados = []

    for registro in registros:
        cliente = normalizar_sem_acento(
            registro.get(coluna_cliente)
        )
        cidade = normalizar_sem_acento(
            registro.get(coluna_cidade)
        )
        unidade = normalizar_sem_acento(
            registro.get(coluna_unidade)
        )
        ctrc = normalizar_texto(
            registro.get(coluna_ctrc)
        )

        if cliente not in clientes_normalizados:
            continue

        if cidade not in cidades_normalizadas:
            continue

        if unidade != unidade_normalizada:
            continue

        if not ctrc:
            continue

        dados_ctrc = decompor_ctrc(ctrc)

        filtrados.append({
            "ctrc_original": ctrc,
            "serie": dados_ctrc["serie"],
            "numero": dados_ctrc["numero"],
            "digito": dados_ctrc["digito"],
            "cliente_pagador": normalizar_texto(
                registro.get(coluna_cliente)
            ),
            "cidade_destinatario": normalizar_texto(
                registro.get(coluna_cidade)
            ),
            "unidade_emissora": normalizar_texto(
                registro.get(coluna_unidade)
            ),
            "registro_original": registro,
        })

    return filtrados

PADRAO_CTRC = re.compile(
    r"^\s*([A-Z]{2,4})\s*(\d+)(?:-(\d+))?\s*$",
    re.IGNORECASE,
)


def decompor_ctrc(
    valor: str,
) -> dict:
    texto = normalizar_texto(valor)

    match = PADRAO_CTRC.match(texto)

    if not match:
        raise ValueError(
            f"Formato de CTRC não reconhecido: {valor}"
        )

    return {
        "serie": match.group(1).upper(),
        "numero": match.group(2),
        "digito": match.group(3) or "",
        "original": texto,
    }

def diagnosticar_filtros(
    registros: list[dict],
    clientes_permitidos: set[str],
    cidades_permitidas: set[str],
    unidade_emissora: str,
) -> dict:
    if not registros:
        return {
            "total": 0,
            "unidade": 0,
            "cidade": 0,
            "cliente": 0,
            "todos_filtros": 0,
            "clientes_encontrados": [],
            "cidades_encontradas": [],
            "unidades_encontradas": [],
        }

    exemplo = registros[0]

    coluna_cliente = encontrar_coluna(
        exemplo,
        CLIENTE_COLUNA,
    )
    coluna_cidade = encontrar_coluna(
        exemplo,
        CIDADE_COLUNA,
    )
    coluna_unidade = encontrar_coluna(
        exemplo,
        UNIDADE_COLUNA,
    )

    clientes_normalizados = {
        normalizar_sem_acento(cliente)
        for cliente in clientes_permitidos
    }

    cidades_normalizadas = {
        normalizar_sem_acento(cidade)
        for cidade in cidades_permitidas
    }

    unidade_normalizada = normalizar_sem_acento(
        unidade_emissora
    )

    total_unidade = 0
    total_cidade = 0
    total_cliente = 0
    total_completo = 0

    clientes_encontrados = set()
    cidades_encontradas = set()
    unidades_encontradas = set()

    for registro in registros:
        cliente_original = normalizar_texto(
            registro.get(coluna_cliente)
        )
        cidade_original = normalizar_texto(
            registro.get(coluna_cidade)
        )
        unidade_original = normalizar_texto(
            registro.get(coluna_unidade)
        )

        cliente = normalizar_sem_acento(
            cliente_original
        )
        cidade = normalizar_sem_acento(
            cidade_original
        )
        unidade = normalizar_sem_acento(
            unidade_original
        )

        if cliente_original:
            clientes_encontrados.add(
                cliente_original
            )

        if cidade_original:
            cidades_encontradas.add(
                cidade_original
            )

        if unidade_original:
            unidades_encontradas.add(
                unidade_original
            )

        atende_unidade = (
            unidade == unidade_normalizada
        )
        atende_cidade = (
            cidade in cidades_normalizadas
        )
        atende_cliente = (
            cliente in clientes_normalizados
        )

        if atende_unidade:
            total_unidade += 1

        if atende_cidade:
            total_cidade += 1

        if atende_cliente:
            total_cliente += 1

        if (
            atende_unidade
            and atende_cidade
            and atende_cliente
        ):
            total_completo += 1

    return {
        "total": len(registros),
        "unidade": total_unidade,
        "cidade": total_cidade,
        "cliente": total_cliente,
        "todos_filtros": total_completo,
        "clientes_encontrados": sorted(
            clientes_encontrados
        ),
        "cidades_encontradas": sorted(
            cidades_encontradas
        ),
        "unidades_encontradas": sorted(
            unidades_encontradas
        ),
    }