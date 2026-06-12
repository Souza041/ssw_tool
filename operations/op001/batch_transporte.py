from pathlib import Path

import pandas as pd

from openpyxl import load_workbook

import unicodedata

from operations.op001.coleta import OP001Coleta

from web.jobs import add_log, set_progress


MAPA_COLUNAS = {
    "CNPJ DESTINATARIO": "CNPJ_DESTINATARIO",
    "NUMERO DO TRANSPORTE": "TRANSPORTE",
    "ORDEM DE VENDA / ORDEM INVERSA": "ORDEM_INVERSA",
    "CLIENTE": "CLIENTE",
    "MUNICIPIO DESTINO": "MUNICIPIO_DESTINO",
    "UF DESTINO": "UF_DESTINO",
    "NOTA FISCAL": "NOTA_FISCAL",
    "SKU": "SKU",
}

COLUNAS_OBRIGATORIAS = {
    "CLIENTE",
    "MUNICIPIO_DESTINO",
    "UF_DESTINO",
    "TRANSPORTE",
    "ORDEM_INVERSA",
    "CNPJ_DESTINATARIO",
}

COLUNA_RESULTADO_COLETA = "COLETA_GERADA"
COLUNA_RESULTADO_SEQ = "SEQ_COLETA"
COLUNA_RESULTADO_STATUS = "STATUS_BOT"
COLUNA_RESULTADO_MSG = "MENSAGEM_BOT"

COLUNA_COLETA_EXISTENTE = "UNNAMED: 3"
COLUNA_TIPO_COLETA = "UNNAMED: 4"

TIPOS_IGNORAR = {
    "RECUSA",
}


def normalizar_texto(texto: str) -> str:
    texto = str(texto).strip().upper()

    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ASCII", "ignore").decode("ASCII")

    return texto


def normalizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df.columns = [
        normalizar_texto(col)
        for col in df.columns
    ]

    return df


def validar_colunas(df: pd.DataFrame) -> None:
    ausentes = COLUNAS_OBRIGATORIAS - set(df.columns)

    if ausentes:
        raise ValueError(
            f"Colunas obrigatórias ausentes: {', '.join(sorted(ausentes))}"
        )


def processar_planilha_transporte(
    op001: OP001Coleta,
    input_file: Path,
    output_file: Path,
    job=None,
) -> Path:
    df = pd.read_excel(input_file, dtype=str).fillna("")
    df = normalizar_colunas(df)
    df = df.rename(columns=MAPA_COLUNAS)

    validar_colunas(df)

    wb = load_workbook(input_file)
    ws = wb.active

    output_file.parent.mkdir(parents=True, exist_ok=True)

    total = len(df)

    if job:
        set_progress(job, 0, total)

    # Cabeçalhos de status no final da planilha
    ws["AG1"] = "STATUS_BOT"
    ws["AH1"] = "MENSAGEM_BOT"

    for index, row in df.iterrows():
        linha_excel = index + 2  # linha 1 é cabeçalho

        coleta_existente = str(ws[f"D{linha_excel}"].value or "").strip()

        if coleta_existente:
            ws[f"AG{linha_excel}"] = "IGNORADO"
            ws[f"AH{linha_excel}"] = f"Linha ignorada. Coleta já existente: {coleta_existente}"

            if job:
                add_log(job, f"Linha {index + 1}/{total} ignorada | coleta já existente={coleta_existente}")
                set_progress(job, index + 1, total)

            wb.save(output_file)
            continue

        tipo_coleta = str(row.get(COLUNA_TIPO_COLETA, "")).strip().upper()

        if tipo_coleta in TIPOS_IGNORAR:
            ws[f"AG{linha_excel}"] = "IGNORADO"
            ws[f"AH{linha_excel}"] = f"Tipo de coleta ignorado: {tipo_coleta}"

            if job:
                add_log(job, f"Linha {index + 1}/{total} ignorada | tipo={tipo_coleta}")
                set_progress(job, index + 1, total)

            wb.save(output_file)
            continue

        try:
            if job:
                add_log(
                    job,
                    f"Processando linha {index + 1}/{total} | "
                    f"Cliente={row['CLIENTE']} | "
                    f"Destino={row['MUNICIPIO_DESTINO']}/{row['UF_DESTINO']}"
                )

            resultado = op001.salvar_coleta_transporte(
                nome_cliente=str(row["CLIENTE"]).strip(),
                municipio_destino=str(row["MUNICIPIO_DESTINO"]).strip(),
                uf_destino=str(row["UF_DESTINO"]).strip(),
                transporte=str(row["TRANSPORTE"]).strip(),
                ordem_inversa=str(row["ORDEM_INVERSA"]).strip(),
                cnpj_destinatario=str(row["CNPJ_DESTINATARIO"]).strip(),
            )

            coleta_gerada = resultado.get("coleta", "")

            # Coluna D recebe o número da coleta
            ws[f"D{linha_excel}"] = coleta_gerada

            # Colunas AG/AH recebem status e mensagem
            ws[f"AG{linha_excel}"] = "OK" if resultado.get("sucesso") else "ERRO"
            ws[f"AH{linha_excel}"] = resultado.get("mensagem", "")

            if job:
                add_log(job, f"Coleta gerada: {coleta_gerada}")

        except Exception as exc:
            ws[f"AG{linha_excel}"] = "ERRO"
            ws[f"AH{linha_excel}"] = str(exc)

            if job:
                add_log(job, f"Erro na linha {index + 1}: {exc}")

        if job:
            set_progress(job, index + 1, total)

        wb.save(output_file)

    wb.save(output_file)

    return output_file