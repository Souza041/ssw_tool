from pathlib import Path

from openpyxl import load_workbook

from operations.op101.comprovantes import OP101Comprovantes


COLUNA_NF = "B"
COLUNA_DATA_OC = "E"
COLUNA_RESULTADO = "F"


def normalizar_data_excel(valor) -> str:
    if not valor:
        return ""

    if hasattr(valor, "strftime"):
        return valor.strftime("%d%m%y")

    texto = str(valor).strip()

    if "/" in texto:
        partes = texto.split("/")
        if len(partes) == 3:
            dia = partes[0].zfill(2)
            mes = partes[1].zfill(2)
            ano = partes[2][-2:]
            return f"{dia}{mes}{ano}"

    texto = "".join(ch for ch in texto if ch.isdigit())

    if len(texto) == 8:
        return texto[:4] + texto[4:6] + texto[6:]

    return texto


def processar_planilha_comprovantes(
    op101: OP101Comprovantes,
    input_file: Path,
    output_file: Path,
    data_ini: str,
    data_fin: str,
    job=None,
) -> Path:
    wb = load_workbook(input_file)
    ws = wb.active

    output_file.parent.mkdir(parents=True, exist_ok=True)

    total = ws.max_row - 1

    if job:
        from web.jobs import set_progress
        set_progress(job, 0, total)

    for row in range(2, ws.max_row + 1):
        numero_nf = str(ws[f"{COLUNA_NF}{row}"].value or "").strip()

        if not numero_nf:
            continue

        try:
            if job:
                from web.jobs import add_log
                add_log(job, f"Consultando linha {row - 1}/{total} | NF={numero_nf}")

            resultado = op101.consultar_comprovante_nf(
                numero_nf=numero_nf,
                data_ini=data_ini,
                data_fin=data_fin,
            )

            ws[f"{COLUNA_RESULTADO}{row}"] = resultado

            if job:
                from web.jobs import add_log
                add_log(job, f"NF={numero_nf} | {resultado}")

        except Exception as exc:
            ws[f"{COLUNA_RESULTADO}{row}"] = f"Erro: {exc}"

            if job:
                from web.jobs import add_log
                add_log(job, f"Erro linha {row - 1} | NF={numero_nf} | {exc}")

        if job:
            from web.jobs import set_progress
            set_progress(job, row - 1, total)

        wb.save(output_file)

    wb.save(output_file)
    return output_file