import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from tools.renomeador.processor import extrair_nf_da_imagem


EXTENSOES = {".png", ".jpg", ".jpeg"}
TEMPLATE_CARRIER = Path("tools/renomeador/templates/Carrier.xltm")

EMAIL_FIXO = "gce4@rodobrastransp.com.br"
LIMITE_POR_ARQUIVO = 20


def nome_unico(pasta: Path, nome_base: str, ext: str) -> Path:
    destino = pasta / f"{nome_base}{ext}"
    contador = 2

    while destino.exists():
        destino = pasta / f"{nome_base}_{contador}{ext}"
        contador += 1

    return destino


def ler_base(base_path: Path) -> pd.DataFrame:
    extensao = base_path.suffix.lower()

    if extensao in {".xlsx", ".xlsm"}:
        abas = pd.read_excel(
            base_path,
            sheet_name=None,
            dtype=str,
            engine="openpyxl",
        )

        abas_normalizadas = {
            str(nome).strip().upper(): df
            for nome, df in abas.items()
        }

        # Ordem de prioridade:
        # primeiro DATA-SP/DATA-SC, depois SP/SC
        nomes_prioritarios = [
            "DATA-SP",
            "DATA-SC",
            "SP",
            "SC",
        ]

        dataframes = []

        for nome_aba in nomes_prioritarios:
            df = abas_normalizadas.get(nome_aba)

            if df is None or df.empty:
                continue

            df = df.copy()
            df["__aba_origem"] = nome_aba
            df["__prioridade"] = (
                0 if nome_aba.startswith("DATA-") else 1
            )

            dataframes.append(df)

        # Caso o arquivo não tenha nenhuma das quatro abas esperadas,
        # usa a primeira aba disponível.
        if not dataframes:
            primeira_aba, primeiro_df = next(iter(abas.items()))

            primeiro_df = primeiro_df.copy()
            primeiro_df["__aba_origem"] = str(primeira_aba).strip()
            primeiro_df["__prioridade"] = 2

            dataframes.append(primeiro_df)

        return pd.concat(
            dataframes,
            ignore_index=True,
            sort=False,
        )

    if extensao == ".xls":
        # Arquivos .xls antigos podem exigir:
        # python -m pip install xlrd
        abas = pd.read_excel(
            base_path,
            sheet_name=None,
            dtype=str,
        )

        dataframes = []

        for nome_aba, df in abas.items():
            nome_normalizado = str(nome_aba).strip().upper()

            if nome_normalizado not in {
                "DATA-SP",
                "DATA-SC",
                "SP",
                "SC",
            }:
                continue

            df = df.copy()
            df["__aba_origem"] = nome_normalizado
            df["__prioridade"] = (
                0 if nome_normalizado.startswith("DATA-") else 1
            )

            dataframes.append(df)

        if not dataframes:
            primeira_aba, primeiro_df = next(iter(abas.items()))

            primeiro_df = primeiro_df.copy()
            primeiro_df["__aba_origem"] = str(primeira_aba).strip()
            primeiro_df["__prioridade"] = 2

            dataframes.append(primeiro_df)

        return pd.concat(
            dataframes,
            ignore_index=True,
            sort=False,
        )

    if extensao != ".csv":
        raise ValueError(
            f"Formato não suportado: {extensao}. "
            "Envie CSV, XLSX, XLSM ou XLS."
        )

    tentativas = [
        {
            "encoding": "utf-16",
            "sep": "\t",
            "engine": "python",
        },
        {
            "encoding": "utf-8-sig",
            "sep": None,
            "engine": "python",
        },
        {
            "encoding": "utf-8",
            "sep": None,
            "engine": "python",
        },
        {
            "encoding": "latin1",
            "sep": None,
            "engine": "python",
        },
        {
            "encoding": "utf-8-sig",
            "sep": ";",
            "engine": "python",
        },
        {
            "encoding": "utf-8-sig",
            "sep": ",",
            "engine": "python",
        },
        {
            "encoding": "latin1",
            "sep": ";",
            "engine": "python",
        },
        {
            "encoding": "latin1",
            "sep": ",",
            "engine": "python",
        },
    ]

    for config in tentativas:
        try:
            df = pd.read_csv(
                base_path,
                dtype=str,
                **config,
            )

            if len(df.columns) > 1:
                df["__aba_origem"] = "CSV"
                df["__prioridade"] = 0
                return df

        except Exception:
            continue

    raise ValueError(
        "Não foi possível identificar a codificação ou o separador "
        "do CSV."
    )


def limpar_texto_base(valor) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip()

    if texto.lower() in {"nan", "none", "nat"}:
        return ""

    # Corrige valores importados como 123456.0
    if texto.endswith(".0"):
        texto = texto[:-2]

    return texto


def converter_data_base(valor):
    texto = limpar_texto_base(valor)

    if not texto:
        return None

    formatos = [
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(
                texto,
                formato,
            ).date()
        except ValueError:
            continue

    try:
        ts = pd.to_datetime(
            texto,
            errors="coerce",
        )

        if not pd.isna(ts):
            return ts.date()

    except Exception:
        pass

    return None


def ler_base_csv(base_csv: Path) -> dict:
    df = ler_base(base_csv)

    df.columns = [
        str(coluna).strip()
        for coluna in df.columns
    ]

    colunas_obrigatorias = {
        "NF No",
        "NS No",
        "Load ID",
    }

    faltantes = colunas_obrigatorias.difference(df.columns)

    if faltantes:
        raise ValueError(
            "A base não possui as colunas obrigatórias: "
            + ", ".join(sorted(faltantes))
        )

    # DATA-SP e DATA-SC vêm primeiro.
    if "__prioridade" in df.columns:
        df = df.sort_values(
            by="__prioridade",
            ascending=True,
            kind="stable",
        )

    mapa = {}

    for _, row in df.iterrows():
        nf = limpar_texto_base(
            row.get("NF No", "")
        ).lstrip("0")

        serie = limpar_texto_base(
            row.get("NS No", "")
        ).lstrip("0")

        load_id = limpar_texto_base(
            row.get("Load ID", "")
        )

        origem = limpar_texto_base(
            row.get("__aba_origem", "")
        )

        if not nf:
            continue

        registered = row.get(
            "Registered Date",
            "",
        )

        data_registro = converter_data_base(
            registered
        )

        novo_registro = {
            "nf": nf,
            "serie": serie,
            "load_id": load_id,
            "occurrence_date": data_registro,
            "truck": limpar_texto_base(
                row.get("Truck", "")
            ),
            "occurrence": limpar_texto_base(
                row.get("Occurrence", "")
            ),
            "approval": limpar_texto_base(
                row.get("Approval", "")
            ),
            "shipto_name": limpar_texto_base(
                row.get("ShipTo Name", "")
            ),
            "protocolos": limpar_texto_base(
                row.get("Protocolos", "")
            ),
            "aba_origem": origem,
        }

        registro_existente = mapa.get(nf)

        if not registro_existente:
            mapa[nf] = novo_registro
            continue

        # Caso a primeira aba não tenha algum valor,
        # complementa usando SP/SC ou outra aba.
        for chave, valor in novo_registro.items():
            atual = registro_existente.get(chave)

            if atual in {None, ""} and valor not in {None, ""}:
                registro_existente[chave] = valor

    return mapa


def transportadora_por_serie(
    serie: str,
    aba_origem: str = "",
) -> str:
    serie_normalizada = (
        str(serie)
        .strip()
        .lstrip("0")
    )

    origem = (
        str(aba_origem)
        .strip()
        .upper()
    )

    if serie_normalizada in {"1", "3", "13"}:
        return "ROBR_SP"

    if serie_normalizada == "33":
        return "ROBR_SC"

    if origem in {"SP", "DATA-SP"}:
        return "ROBR_SP"

    if origem in {"SC", "DATA-SC"}:
        return "ROBR_SC"

    return "ROBR_SC"


def gerar_protocolos_carrier(
    registros: list[dict],
    output_dir: Path,
    email: str,
    observacao: str,
    job=None,
) -> list[Path]:
    arquivos = []
    grupos = {}

    for item in registros:
        serie = str(item.get("serie") or "").strip()
        aba_origem = str(item.get("aba_origem") or "").strip()

        if serie or aba_origem:
            grupo = transportadora_por_serie(serie, aba_origem)
        else:
            grupo = "SEM_BASE"

        grupos.setdefault(grupo, []).append(item)

    for transportadora, itens in grupos.items():
        for idx in range(0, len(itens), LIMITE_POR_ARQUIVO):
            lote = itens[idx:idx + LIMITE_POR_ARQUIVO]
            numero_lote = (
                idx // LIMITE_POR_ARQUIVO
            ) + 1

            wb = load_workbook(
                TEMPLATE_CARRIER,
                keep_vba=True,
            )

            ws = wb.active

            protocolo = gerar_numero_protocolo()

            ws["C1"] = (
                ""
                if transportadora == "SEM_BASE"
                else transportadora
            )
            ws["C2"] = datetime.now()
            ws["C2"].number_format = "DD/MM/YYYY"

            ws["C3"] = protocolo
            ws["C4"] = email

            for linha in range(9, 80):
                for coluna in range(2, 7):
                    ws.cell(
                        linha,
                        coluna,
                    ).value = None

            linha_excel = 9

            for item in lote:
                load_id = str(item.get("load_id") or "").strip()
                nf = str(item.get("nf") or "").strip()
                serie = str(item.get("serie") or "").strip()
                data_assinatura = item.get("data_assinatura")

                ws.cell(linha_excel, 2).value = (
                    int(load_id)
                    if load_id.isdigit()
                    else None
                )

                ws.cell(linha_excel, 3).value = (
                    int(nf)
                    if nf.isdigit()
                    else nf
                )

                ws.cell(linha_excel, 4).value = (
                    int(serie)
                    if serie.isdigit()
                    else None
                )

                ws.cell(linha_excel, 5).value = data_assinatura
                ws.cell(linha_excel, 5).number_format = "DD/MM/YYYY"

                ws.cell(linha_excel, 6).value = observacao

                linha_excel += 1

            nome = f"{protocolo}.xltm"
            caminho = output_dir / nome

            while caminho.exists():
                protocolo = gerar_numero_protocolo()
                nome = f"{protocolo}.xltm"
                caminho = output_dir / nome
                ws["C3"] = protocolo

            wb.save(caminho)

            arquivos.append(caminho)

            if job:
                job.logs.put(
                    f"Carrier gerado: {nome} "
                    f"({len(lote)} notas)"
                )

    return arquivos


def adicionar_zip(
    zipf,
    arquivo: Path,
    base: Path,
):
    if arquivo.is_file():
        zipf.write(
            arquivo,
            arquivo.relative_to(base),
        )

def gerar_numero_protocolo() -> str:
    agora = datetime.now()

    return (
        agora.strftime("%Y%m%d%H%M%S")
        + f"{agora.microsecond // 1000:03d}"
    )

def processar_carrier_lg(
    input_dir: Path,
    base_csv: Path,
    output_dir: Path,
    zip_path: Path,
    email: str = EMAIL_FIXO,
    observacao: str = "",
    job=None,
) -> Path:
    renomeados_dir = output_dir / "renomeados"
    falhou_dir = output_dir / "falhou"
    sem_base_dir = output_dir / "sem_base"

    renomeados_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    falhou_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    sem_base_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    if job:
        job.logs.put("Lendo a base diária...")

    mapa_base = ler_base_csv(base_csv)

    if job:
        job.logs.put(
            f"Base carregada: {len(mapa_base)} NFs."
        )

    arquivos = sorted(
        [
            p
            for p in input_dir.iterdir()
            if (
                p.is_file()
                and p.suffix.lower() in EXTENSOES
            )
        ],
        key=lambda item: item.name.lower(),
    )

    if job:
        job.progress = 0
        job.total = len(arquivos)

    registros_carrier = []
    total_ocr = 0
    total_nome = 0
    total_falhou = 0
    total_sem_base = 0

    for idx, arquivo in enumerate(
        arquivos,
        start=1,
    ):
        if job:
            job.logs.put(
                f"Processando {idx}/{len(arquivos)}: "
                f"{arquivo.name}"
            )

        nome = arquivo.stem.strip()

        # Aceita NF já presente no nome:
        # 117824.png, 000117824.png etc.
        if re.fullmatch(r"\d{5,9}", nome):
            nf = nome.lstrip("0")
            metodo = "nome do arquivo"
            total_nome += 1

            if job:
                job.logs.put(
                    f"NF obtida pelo nome: {arquivo.name}"
                )

        else:
            if job:
                job.logs.put(
                    f"Iniciando OCR: {arquivo.name}"
                )

            nf, _, metodo = extrair_nf_da_imagem(
                arquivo
            )

            total_ocr += 1

        if not nf:
            destino = nome_unico(
                falhou_dir,
                arquivo.stem,
                arquivo.suffix.lower(),
            )

            shutil.copy2(
                arquivo,
                destino,
            )

            total_falhou += 1

            if job:
                job.logs.put(
                    f"Falhou OCR: {arquivo.name}"
                )
                job.progress = idx

            continue

        nf = str(nf).strip().lstrip("0")

        destino = nome_unico(
            renomeados_dir,
            nf,
            arquivo.suffix.lower(),
        )

        shutil.copy2(
            arquivo,
            destino,
        )

        dados_base = mapa_base.get(nf)

        if not dados_base:
            destino_sem_base = nome_unico(
                sem_base_dir,
                nf,
                arquivo.suffix.lower(),
            )

            shutil.copy2(
                arquivo,
                destino_sem_base,
            )

            total_sem_base += 1

            # Inclui a NF em um Carrier separado de auditoria.
            registros_carrier.append({
                "nf": nf,
                "serie": "",
                "load_id": "",
                "data_assinatura": None,
                "aba_origem": "",
            })

            if job:
                job.logs.put(
                    f"Sem base: NF {nf} será incluída "
                    f"na planilha de auditoria."
                )
                job.progress = idx

            continue

        registros_carrier.append({
            "nf": nf,
            "serie": dados_base["serie"],
            "load_id": dados_base["load_id"],
            "data_assinatura": dados_base[
                "occurrence_date"
            ],
            "aba_origem": dados_base.get("aba_origem", ""),
        })

        if job:
            data_log = dados_base.get(
                "occurrence_date"
            )

            data_log = (
                data_log.strftime("%d/%m/%Y")
                if data_log
                else "sem data"
            )

            job.logs.put(
                f"OK: {arquivo.name} -> {destino.name} | "
                f"Load {dados_base['load_id']} | "
                f"Série {dados_base['serie']} | "
                f"Data {data_log} | "
                f"{metodo}"
            )

            job.progress = idx

    carrier_dir = output_dir / "carrier"
    carrier_dir.mkdir(exist_ok=True)

    gerar_protocolos_carrier(
        registros=registros_carrier,
        output_dir=carrier_dir,
        email=email,
        observacao=observacao,
        job=job,
    )

    if job:
        job.logs.put("Compactando resultado...")

    with zipfile.ZipFile(
        zip_path,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as zipf:
        for arquivo in output_dir.rglob("*"):
            if arquivo.is_file():
                adicionar_zip(
                    zipf,
                    arquivo,
                    output_dir,
                )

    if job:
        job.logs.put(
            f"Identificadas pelo nome: {total_nome}"
        )
        job.logs.put(
            f"Enviadas ao OCR: {total_ocr}"
        )
        job.logs.put(
            f"Falhas no OCR: {total_falhou}"
        )
        job.logs.put(
            f"Sem correspondência na base: {total_sem_base}"
        )
        job.logs.put(
            f"Notas adicionadas ao Carrier: "
            f"{len(registros_carrier)}"
        )
        job.logs.put("ZIP final gerado.")

    return zip_path